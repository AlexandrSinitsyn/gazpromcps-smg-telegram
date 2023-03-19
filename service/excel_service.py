import csv
import openpyxl
from openpyxl.utils import get_column_letter
from typing import List, Dict

from repository.job_repository import *

path_to_done = '/bot/storage/done/'
path_to_job_list = '/bot/storage/jobs/'

os.makedirs(path_to_done, exist_ok=True)
os.makedirs(path_to_job_list, exist_ok=True)


class ExcelReader:
    _workbook: openpyxl.Workbook
    _blue = 'FFB9CDE5'
    _green = 'FF33CC33'

    def __init__(self, filename: str):
        self._workbook = openpyxl.load_workbook(filename)

    def table(self):
        return {sheet: [list(row) for row in self._workbook[sheet].values] for sheet in self._workbook.sheetnames}

    def get_color(self, sheet: str, row: int, col: int):
        return self._workbook[sheet][get_column_letter(col + 1) + str(row)].fill.start_color.index

    def is_green(self, sheet: str, row: int, col: int):
        return self.get_color(sheet, row, col) == self._green

    def is_blue(self, sheet: str, row: int, col: int):
        return self.get_color(sheet, row, col) in [self._blue, '3', '4']


def write_xlsx(file_name: str, data: Dict[str, List[List[str]]]):
    workbook = openpyxl.Workbook()

    workbook.remove(workbook['Sheet'])

    for name, rows in data.items():
        sheet = workbook.create_sheet(name)
        for row in rows:
            sheet.append(row)

    workbook.save(file_name)


class ExcelService:
    @staticmethod
    def import_data(data: List[Job]):
        for j in data:
            save_job(j, find_by_params(j.stage, j.master, j.title) is None)

    @staticmethod
    def import_csv(file_name: str):
        data = []

        def capitalize_first(text: str):
            return text[0].upper() + text[1:]

        with open(path_to_job_list + file_name, 'r') as f:
            r = csv.reader(f, delimiter=',')

            for row in r:
                data.append(Job(-1, row[1], row[2], capitalize_first(row[0]), row[3], True, datetime.now()))

        ExcelService.import_data(data)

    @staticmethod
    def rollback(timestamp: datetime):
        ExcelService.import_csv('save_' + str(timestamp) + '.csv')

    @staticmethod
    def delete_all():
        drop_table()

    @staticmethod
    def summarise(start: datetime) -> List[str]:
        total = collect_daily(start)

        res = {}
        for cj in total:
            if cj.job not in res:
                res[cj.job] = 0
            res[cj.job] += cj.count

        return [str(CompletedJob(-1, k, v, datetime.now()))for k, v in res.items()]

    @staticmethod
    def export_csv(start: datetime) -> str:
        return '\n'.join([CompletedJob.csv_title()] + [str(cj) for cj in collect_daily(start)] +\
                         [] + ['Сумма по каждому виду работ'] + ExcelService.summarise(start))

    @staticmethod
    def export_xlsx(start: datetime) -> List[List[str]]:
        return [[e.replace('"', '') for e in CompletedJob.csv_title().split('","')]] +\
               [[e.replace('"', '') for e in str(cj).split('","')] for cj in collect_daily(start)] +\
               [[]] + [['Сумма по каждому виду работ']] +\
               [[e.replace('"', '') for e in str(cj).split('","')] for cj in ExcelService.summarise(start)]

    @staticmethod
    def save(start: datetime, xlsx: bool = False) -> str:
        now = datetime.now()

        file_name = path_to_done + 'save_' + datetime.strftime(now, '%Y%m%d_%H%M%S') + ('.xlsx' if xlsx else '.csv')

        if xlsx:
            write_xlsx(file_name,
                       {f"Отчет от {datetime.strftime(now, '%Y-%m-%d (%H-%M)')}": ExcelService.export_xlsx(start)})
        else:
            with open(file_name, 'w', newline='', encoding='utf-8') as f:
                f.write(ExcelService.export_csv(start))

        return file_name
